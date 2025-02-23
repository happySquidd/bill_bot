from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime, timedelta


def create_xlsx(excel_path, spectrum_bill_date, spectrum_from_date, spectrum_to_date, spectrum_total, duke_bill_date, duke_from_date, duke_to_date, duke_total):
    todays_date = datetime.now()
    current_month = todays_date.strftime("%B")
    current_day = todays_date.day
    current_year = todays_date.year
    new_date = todays_date + timedelta(11)
    pay_by_day = new_date.strftime("%d")
    pay_by_month = new_date.strftime("%B")
    pay_by_year = new_date.strftime("%Y")
    
    wb = Workbook()
    ws = wb.active

    matrix_letlers = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]

    ws['A2'] = "Prepared and sent:" #blue
    ws['B2'] = str(current_month) + " " + str(current_day) + ", " + str(current_year) #blue

    ws['A4'] = "Spectrum" #bold
    ws['G4'] = "Duke Energy" #bold

    ### spectrum ###
    ws['A6'] = "Bill Date" #bold
    ws['B6'] = "From" #bold
    ws['C6'] = "Through" #bold
    ws['D6'] = "Amount" #bold
    ws['E6'] = "50%" # gray-fill #bold

    ws['A7'] = str(spectrum_bill_date) #bold
    ws['B7'] = str(spectrum_from_date)
    ws['C7'] = str(spectrum_to_date)
    ws['D7'] = "$ " + str(spectrum_total)
    ws['E7'] = str(round(spectrum_total/2, 2)) #gray-fill

    ws['E8'] = "" #gray-fill

    ws['C9'] =  "Subtotal"
    ws['D9'] = "$ " + str(spectrum_total)
    ws['E9'] = "$ " + str(round(spectrum_total/2, 2)) #gray-fill


    ### duke energy ###
    ws['G6'] = "Bill Date" #bold
    ws['H6'] = "From" #bold
    ws['I6'] = "Through" #bold
    ws['J6'] = "Amount" #bold
    ws['K6'] = "50%" # gray-fill

    ws['G7'] = str(duke_bill_date) #bold
    ws['H7'] = str(duke_from_date)
    ws['I7'] = str(duke_to_date)
    ws['J7'] = "$ " + str(duke_total)
    ws['K7'] = "$ " + str(round(duke_total/2, 2)) #gray-fill

    ws['K8'] = "" #gray-fill

    ws['I9'] =  "Subtotal"
    ws['J9'] = "$ " + str(duke_total)
    ws['K9'] = "$ " + str(round(duke_total/2, 2)) #gray-fill
    ### duke ###

    ### totals ###
    ws['J13'] = "Total amount:" # bold
    ws['K13'] = "$ " + str(round(duke_total/2+spectrum_total/2, 2)) #bold #highlight

    ws['J14'] = "Please pay by:" # bold
    ws['K14'] = str(pay_by_month) + " " + str(pay_by_day) + ", " + str(pay_by_year) # bold #highlight
    ###

    ### styling ###
    for letter in matrix_letlers:
        ws.column_dimensions[letter].width = 20

    header_style = Font(bold=True, color="4f81bd")
    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
    yellow_fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')

    bold_font_matrix = ['A4', 'G4', 'A6', 'B6', 'C6', 'D6', 'E6', 'G6', 'H6', 'I6', 'J6', 'K6', 'A7', 'G7', 'C9', 'E9', 'I9', 'K9', 'J13', 'K13', 'J14', 'K14']

    for box in bold_font_matrix:
        ws[box].font = bold_font

    ws['A2'].font = header_style
    ws['B2'].font = header_style

    grey_fill_matrix = ['E6', 'E7', 'E8', 'E9', 'K6', 'K7', 'K8', 'K9', 'J13', 'J14']

    for box in grey_fill_matrix:
        ws[box].fill = grey_fill

    ws['K13'].fill = yellow_fill
    ws['K14'].fill = yellow_fill


    thin_border_top = Border(top=Side(style='thin'))

    thin_border_top_matrix = ['A7', 'B7', 'C7', 'D7', 'E7', 'G7', 'H7', 'I7', 'J7', 'K7', 'C9', 'D9', 'E9', 'I9', 'J9', 'K9']

    for box in thin_border_top_matrix:
        ws[box].border = thin_border_top

    ws['J13'].border = Border(top=Side(style='thick'),
                            left=Side(style='thick'))
    ws['K13'].border = Border(top=Side(style='thick'),
                            right=Side(style='thick'))
    ws['J14'].border = Border(bottom=Side(style='thick'),
                            left=Side(style='thick'))
    ws['K14'].border = Border(right=Side(style='thick'),
                            bottom=Side(style='thick'))


    wb.save(excel_path)